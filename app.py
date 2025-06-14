
from flask import Flask, render_template, request, send_file
from openpyxl import load_workbook
from io import BytesIO
import datetime

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('form.html')

@app.route('/gerar', methods=['POST'])
def gerar():
    dados = request.form
    wb = load_workbook('modelo.xlsx')
    ws = wb.active

    ws['B3'] = dados.get('nome_cliente')
    ws['B4'] = dados.get('telefone')
    ws['B6'] = dados.get('data_pedido')
    ws['B7'] = dados.get('entrega')
    ws['B11'] = dados.get('colarinho')
    ws['D11'] = dados.get('tipo_colarinho')
    ws['B12'] = dados.get('torax')
    ws['D12'] = dados.get('tipo_frente')
    ws['B13'] = dados.get('pala')
    ws['D13'] = dados.get('tipo_pense')
    ws['B14'] = dados.get('manga')
    ws['B15'] = dados.get('antebraco')
    ws['B16'] = dados.get('biceps')
    ws['B17'] = dados.get('punho_direito')
    ws['B18'] = dados.get('punho_esquerdo')
    ws['B19'] = dados.get('cintura')
    ws['B20'] = dados.get('quadril')
    ws['B21'] = dados.get('comprimento_total')
    ws['B23'] = dados.get('obs_medidas')
    ws['B27'] = dados.get('colarinho_pedido')
    ws['B28'] = dados.get('punhos')
    ws['B29'] = dados.get('frente')
    ws['B30'] = dados.get('monograma')
    ws['B31'] = dados.get('botoes')
    ws['B32'] = dados.get('obs_pedido')
    ws['B36'] = dados.get('ref_tecido')
    ws['B37'] = dados.get('quantidade')
    ws['B38'] = dados.get('fornecedor')

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    nome_arquivo = f"Ficha_DiRoveris_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(output, download_name=nome_arquivo, as_attachment=True)

if __name__ == '__main__':
    app.run()
